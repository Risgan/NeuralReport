import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.calculator import StoreTaxes


# ─── Estilos ────────────────────────────────────────────────

HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
TOTAL_FILL     = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT     = Font(bold=True, size=11)
CURRENCY_FMT   = '#,##0.00'
BORDER_SIDE    = Side(style="thin", color="BFBFBF")
CELL_BORDER    = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)
CENTER_ALIGN   = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN    = Alignment(horizontal="right",  vertical="center")


def _apply_header(cell, value: str):
    cell.value        = value
    cell.font         = HEADER_FONT
    cell.fill         = HEADER_FILL
    cell.border       = CELL_BORDER
    cell.alignment    = CENTER_ALIGN


def _apply_currency(cell, value: float):
    cell.value        = value
    cell.number_format= CURRENCY_FMT
    cell.border       = CELL_BORDER
    cell.alignment    = RIGHT_ALIGN


def _apply_total(cell, value: float):
    cell.value        = value
    cell.number_format= CURRENCY_FMT
    cell.font         = TOTAL_FONT
    cell.fill         = TOTAL_FILL
    cell.border       = CELL_BORDER
    cell.alignment    = RIGHT_ALIGN


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ─── Horizontal ─────────────────────────────────────────────

def export_horizontal(
    rows:   list[StoreTaxes],
    title:  str,
    month1: str,
    month2: str,
    month3: str,
) -> bytes:
    """
    Genera el Excel horizontal:
    Una fila por tienda con columnas:
    Tienda | Mes1 | Mes2 | Mes3 | Total | IVA | ReteFuente | ReteICA | Total Neto
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe"

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"].value     = title
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = CENTER_ALIGN

    # Encabezados
    headers = [
        "Tienda", month1, month2, month3,
        "Total", "IVA", "ReteFuente", "ReteICA", "Total Neto",
    ]
    for col, header in enumerate(headers, start=1):
        _apply_header(ws.cell(row=2, column=col), header)

    # Datos
    for row_idx, store in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1).value  = store.store_code
        ws.cell(row=row_idx, column=1).border = CELL_BORDER
        ws.cell(row=row_idx, column=1).alignment = CENTER_ALIGN

        _apply_currency(ws.cell(row=row_idx, column=2), store.value_month1)
        _apply_currency(ws.cell(row=row_idx, column=3), store.value_month2)
        _apply_currency(ws.cell(row=row_idx, column=4), store.value_month3)
        _apply_currency(ws.cell(row=row_idx, column=5), store.total_value)
        _apply_currency(ws.cell(row=row_idx, column=6), store.iva)
        _apply_currency(ws.cell(row=row_idx, column=7), store.retefuente)
        _apply_currency(ws.cell(row=row_idx, column=8), store.rete_ica)
        _apply_currency(ws.cell(row=row_idx, column=9), store.total_neto)

    # Fila de totales
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1).value     = "TOTAL"
    ws.cell(row=total_row, column=1).font      = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill      = TOTAL_FILL
    ws.cell(row=total_row, column=1).border    = CELL_BORDER
    ws.cell(row=total_row, column=1).alignment = CENTER_ALIGN

    _apply_total(ws.cell(row=total_row, column=2), sum(r.value_month1 for r in rows))
    _apply_total(ws.cell(row=total_row, column=3), sum(r.value_month2 for r in rows))
    _apply_total(ws.cell(row=total_row, column=4), sum(r.value_month3 for r in rows))
    _apply_total(ws.cell(row=total_row, column=5), sum(r.total_value   for r in rows))
    _apply_total(ws.cell(row=total_row, column=6), sum(r.iva           for r in rows))
    _apply_total(ws.cell(row=total_row, column=7), sum(r.retefuente    for r in rows))
    _apply_total(ws.cell(row=total_row, column=8), sum(r.rete_ica      for r in rows))
    _apply_total(ws.cell(row=total_row, column=9), sum(r.total_neto    for r in rows))

    # Anchos de columna
    widths = [10, 16, 16, 16, 16, 16, 16, 16, 16]
    for col, width in enumerate(widths, start=1):
        _set_col_width(ws, col, width)

    return _to_bytes(wb)


# ─── Vertical ───────────────────────────────────────────────

def export_vertical(
    rows:   list[StoreTaxes],
    title:  str,
    month1: str,
    month2: str,
    month3: str,
) -> bytes:
    """
    Genera el Excel vertical:
    Cada combinación tienda-mes es una fila con columnas:
    Tienda | Mes | Valor | IVA | ReteFuente | ReteICA | Total Neto
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Informe Vertical"

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"].value     = title
    ws["A1"].font      = Font(bold=True, size=13)
    ws["A1"].alignment = CENTER_ALIGN

    # Encabezados
    headers = ["Tienda", "Mes", "Valor", "IVA", "ReteFuente", "ReteICA", "Total Neto"]
    for col, header in enumerate(headers, start=1):
        _apply_header(ws.cell(row=2, column=col), header)

    months_map = {1: month1, 2: month2, 3: month3}
    values_map = {
        1: lambda r: r.value_month1,
        2: lambda r: r.value_month2,
        3: lambda r: r.value_month3,
    }

    current_row = 3
    for store in rows:
        for month_num in range(1, 4):
            value = values_map[month_num](store)

            # Proporcionar impuestos solo en el mes donde tiene valor
            # para no triplicar los totales
            has_value = value > 0
            iva  = round(value * 0.19,   2) if has_value else 0
            fte  = round(value * 0.035,  2) if has_value else 0
            ica  = round(value * 0.0069, 4) if has_value else 0
            neto = round(value - fte - ica, 2) if has_value else 0

            ws.cell(row=current_row, column=1).value     = store.store_code
            ws.cell(row=current_row, column=1).border    = CELL_BORDER
            ws.cell(row=current_row, column=1).alignment = CENTER_ALIGN
            ws.cell(row=current_row, column=2).value     = months_map[month_num]
            ws.cell(row=current_row, column=2).border    = CELL_BORDER
            ws.cell(row=current_row, column=2).alignment = CENTER_ALIGN

            _apply_currency(ws.cell(row=current_row, column=3), value)
            _apply_currency(ws.cell(row=current_row, column=4), iva)
            _apply_currency(ws.cell(row=current_row, column=5), fte)
            _apply_currency(ws.cell(row=current_row, column=6), ica)
            _apply_currency(ws.cell(row=current_row, column=7), neto)

            current_row += 1

    # Anchos de columna
    widths = [10, 14, 16, 16, 16, 16, 16]
    for col, width in enumerate(widths, start=1):
        _set_col_width(ws, col, width)

    return _to_bytes(wb)


# ─── Helper ─────────────────────────────────────────────────

def _to_bytes(wb: Workbook) -> bytes:
    """Convierte el workbook a bytes para enviarlo como respuesta o subirlo a MinIO."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()