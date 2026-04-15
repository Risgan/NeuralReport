import io
from dataclasses import dataclass

from openpyxl import load_workbook


@dataclass
class StoreDay:
    """Licencias de una tienda en un día específico."""
    store_code: int
    day:        int
    licenses:   int


@dataclass
class StoreTotal:
    """Total calculado por tienda en un mes."""
    store_code:     int
    total_days:     int
    total_licenses: int
    total_value:    float


@dataclass
class ParsedFile:
    """Resultado completo de procesar un Excel mensual."""
    importe:  float
    details:  list[StoreDay]
    totals:   list[StoreTotal]


def parse_excel(file_bytes: bytes) -> ParsedFile:
    """
    Lee un Excel mensual y extrae:
    - importe (valor cobro del mes)
    - detalle diario por tienda (file_details)
    - totales por tienda (file_totals)

    Estructura esperada del Excel:
    - Fila 3: contiene 'Valor Cobro :' en col 28 y el importe en col 32
    - Fila 4: encabezados (Tienda, Dia 1 ... Dia 31, Total Dias, Total Valor)
    - Fila 5+: datos por tienda
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Buscar la primera hoja con datos de tiendas
    ws = _find_data_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))

    importe = _extract_importe(rows)
    details, totals = _extract_store_data(rows, importe)

    wb.close()
    return ParsedFile(importe=importe, details=details, totals=totals)


def _find_data_sheet(wb):
    """
    Retorna la hoja que contiene datos de tiendas.
    Busca la que tenga 'Tienda' en la primera columna de la fila 4.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        if rows and rows[0][0] == "Tienda":
            return ws
    # Si no encuentra por nombre, usa la primera hoja
    return wb.active


def _extract_importe(rows: list) -> float:
    """
    Extrae el valor base por licencia desde la fila de configuración.
    Busca 'Valor Cobro :' en cualquier celda de las primeras 5 filas.
    """
    for row in rows[:5]:
        for i, cell in enumerate(row):
            if cell == "Valor Cobro :" and i + 4 < len(row):
                value = row[i + 4]
                if isinstance(value, (int, float)):
                    return float(value)
    # Si no encuentra el valor cobro retorna 0 para no romper el flujo
    return 0.0


def _extract_store_data(
    rows: list,
    importe: float,
) -> tuple[list[StoreDay], list[StoreTotal]]:
    """
    Extrae el detalle diario y los totales por tienda.
    Empieza desde la fila 5 (índice 4) donde están los datos.
    """
    details: list[StoreDay]   = []
    totals:  list[StoreTotal] = []

    for row in rows[4:]:
        # La primera columna es el código de tienda (entero)
        if not row[0] or not isinstance(row[0], (int, float)):
            continue

        store_code = int(row[0])

        # Columnas 1 a 31 son los días (índices 1-31)
        total_days     = 0
        total_licenses = 0

        for day_index in range(1, 32):
            if day_index >= len(row):
                break

            cell_value = row[day_index]
            licenses   = int(cell_value) if isinstance(cell_value, (int, float)) else 0

            details.append(StoreDay(
                store_code=store_code,
                day=day_index,
                licenses=licenses,
            ))

            if licenses > 0:
                total_days     += 1
                total_licenses += licenses

        total_value = round(total_licenses * importe, 2)

        totals.append(StoreTotal(
            store_code=store_code,
            total_days=total_days,
            total_licenses=total_licenses,
            total_value=total_value,
        ))

    return details, totals